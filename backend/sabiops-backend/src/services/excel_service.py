import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from typing import Dict, List, Any, Optional
import logging

logger = logging.getLogger(__name__)

class ExcelService:
    """
    Comprehensive Excel export service for generating business reports and data exports
    """
    
    def __init__(self):
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.right_alignment = Alignment(horizontal='right', vertical='center')
    
    def _apply_header_style(self, worksheet, row_num: int, start_col: int, end_col: int):
        """Apply header styling to a range of cells"""
        for col in range(start_col, end_col + 1):
            cell = worksheet.cell(row=row_num, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border
    
    def _apply_data_style(self, worksheet, start_row: int, end_row: int, start_col: int, end_col: int):
        """Apply data styling to a range of cells"""
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.border = self.border
                cell.alignment = Alignment(vertical='center')
    
    def _auto_adjust_columns(self, worksheet):
        """Auto-adjust column widths based on content"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def export_customers_to_excel(self, customers_data: List[Dict], output_path: str) -> bool:
        """
        Export customers data to Excel
        
        Args:
            customers_data: List of customer data dictionaries
            output_path: Path where Excel file will be saved
        
        Returns:
            Boolean indicating success
        """
        try:
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Customers"
            
            # Add title
            worksheet['A1'] = "CUSTOMER REPORT"
            worksheet['A1'].font = Font(size=16, bold=True)
            worksheet['A1'].alignment = self.center_alignment
            worksheet.merge_cells('A1:H1')
            
            # Add generation date
            worksheet['A2'] = f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
            worksheet['A2'].font = Font(italic=True)
            worksheet.merge_cells('A2:H2')
            
            # Headers
            headers = ['Customer ID', 'Name', 'Email', 'Phone', 'Address', 'Status', 'Created Date', 'Total Revenue']
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=4, column=col, value=header)
            
            self._apply_header_style(worksheet, 4, 1, len(headers))
            
            # Data rows
            for row, customer in enumerate(customers_data, 5):
                worksheet.cell(row=row, column=1, value=customer.get('id', ''))
                worksheet.cell(row=row, column=2, value=customer.get('name', ''))
                worksheet.cell(row=row, column=3, value=customer.get('email', ''))
                worksheet.cell(row=row, column=4, value=customer.get('phone', ''))
                worksheet.cell(row=row, column=5, value=customer.get('address', ''))
                worksheet.cell(row=row, column=6, value=customer.get('status', '').title())
                worksheet.cell(row=row, column=7, value=customer.get('created_at', ''))
                
                # Format revenue as currency
                revenue_cell = worksheet.cell(row=row, column=8, value=customer.get('total_revenue', 0))
                revenue_cell.number_format = '₦#,##0.00'
                revenue_cell.alignment = self.right_alignment
            
            # Apply data styling
            if customers_data:
                self._apply_data_style(worksheet, 5, 4 + len(customers_data), 1, len(headers))
            
            # Summary section
            summary_row = 6 + len(customers_data)
            worksheet.cell(row=summary_row, column=1, value="SUMMARY").font = Font(bold=True)
            worksheet.cell(row=summary_row + 1, column=1, value="Total Customers:")
            worksheet.cell(row=summary_row + 1, column=2, value=len(customers_data))
            
            active_customers = len([c for c in customers_data if c.get('status') == 'active'])
            worksheet.cell(row=summary_row + 2, column=1, value="Active Customers:")
            worksheet.cell(row=summary_row + 2, column=2, value=active_customers)
            
            total_revenue = sum(c.get('total_revenue', 0) for c in customers_data)
            worksheet.cell(row=summary_row + 3, column=1, value="Total Revenue:")
            revenue_summary_cell = worksheet.cell(row=summary_row + 3, column=2, value=total_revenue)
            revenue_summary_cell.number_format = '₦#,##0.00'
            
            self._auto_adjust_columns(worksheet)
            workbook.save(output_path)
            
            logger.info(f"Customer Excel export generated successfully: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Failed to generate customer Excel export: {str(e)}")
            return False
    
    def export_invoices_to_excel(self, invoices_data: List[Dict], output_path: str) -> bool:
        """
        Export invoices data to Excel
        
        Args:
            invoices_data: List of invoice data dictionaries
            output_path: Path where Excel file will be saved
        
        Returns:
            Boolean indicating success
        """
        try:
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Invoices"
            
            # Add title
            worksheet['A1'] = "INVOICE REPORT"
            worksheet['A1'].font = Font(size=16, bold=True)
            worksheet['A1'].alignment = self.center_alignment
            worksheet.merge_cells('A1:I1')
            
            # Add generation date
            worksheet['A2'] = f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
            worksheet['A2'].font = Font(italic=True)
            worksheet.merge_cells('A2:I2')
            
            # Headers
            headers = ['Invoice #', 'Customer', 'Invoice Date', 'Due Date', 'Status', 'Subtotal', 'Tax', 'Total', 'Notes']
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=4, column=col, value=header)
            
            self._apply_header_style(worksheet, 4, 1, len(headers))
            
            # Data rows
            for row, invoice in enumerate(invoices_data, 5):
                worksheet.cell(row=row, column=1, value=invoice.get('invoice_number', ''))
                worksheet.cell(row=row, column=2, value=invoice.get('customer_name', ''))
                worksheet.cell(row=row, column=3, value=invoice.get('invoice_date', ''))
                worksheet.cell(row=row, column=4, value=invoice.get('due_date', ''))
                worksheet.cell(row=row, column=5, value=invoice.get('status', '').title())
                
                # Format currency fields
                subtotal_cell = worksheet.cell(row=row, column=6, value=invoice.get('subtotal', 0))
                subtotal_cell.number_format = '₦#,##0.00'
                subtotal_cell.alignment = self.right_alignment
                
                tax_cell = worksheet.cell(row=row, column=7, value=invoice.get('tax_amount', 0))
                tax_cell.number_format = '₦#,##0.00'
                tax_cell.alignment = self.right_alignment
                
                total_cell = worksheet.cell(row=row, column=8, value=invoice.get('total_amount', 0))
                total_cell.number_format = '₦#,##0.00'
                total_cell.alignment = self.right_alignment
                
                worksheet.cell(row=row, column=9, value=invoice.get('notes', ''))
            
            # Apply data styling
            if invoices_data:
                self._apply_data_style(worksheet, 5, 4 + len(invoices_data), 1, len(headers))
            
            # Summary section
            summary_row = 6 + len(invoices_data)
            worksheet.cell(row=summary_row, column=1, value="SUMMARY").font = Font(bold=True)
            
            total_invoices = len(invoices_data)
            paid_invoices = len([i for i in invoices_data if i.get('status') == 'paid'])
            pending_invoices = len([i for i in invoices_data if i.get('status') == 'pending'])
            overdue_invoices = len([i for i in invoices_data if i.get('status') == 'overdue'])
            
            worksheet.cell(row=summary_row + 1, column=1, value="Total Invoices:")
            worksheet.cell(row=summary_row + 1, column=2, value=total_invoices)
            
            worksheet.cell(row=summary_row + 2, column=1, value="Paid Invoices:")
            worksheet.cell(row=summary_row + 2, column=2, value=paid_invoices)
            
            worksheet.cell(row=summary_row + 3, column=1, value="Pending Invoices:")
            worksheet.cell(row=summary_row + 3, column=2, value=pending_invoices)
            
            worksheet.cell(row=summary_row + 4, column=1, value="Overdue Invoices:")
            worksheet.cell(row=summary_row + 4, column=2, value=overdue_invoices)
            
            total_amount = sum(i.get('total_amount', 0) for i in invoices_data)
            worksheet.cell(row=summary_row + 5, column=1, value="Total Amount:")
            total_summary_cell = worksheet.cell(row=summary_row + 5, column=2, value=total_amount)
            total_summary_cell.number_format = '₦#,##0.00'
            
            self._auto_adjust_columns(worksheet)
            workbook.save(output_path)
            
            logger.info(f"Invoice Excel export generated successfully: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Failed to generate invoice Excel export: {str(e)}")
            return False
    
    def export_products_to_excel(self, products_data: List[Dict], output_path: str) -> bool:
        """
        Export products data to Excel
        
        Args:
            products_data: List of product data dictionaries
            output_path: Path where Excel file will be saved
        
        Returns:
            Boolean indicating success
        """
        try:
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Products"
            
            # Add title
            worksheet['A1'] = "PRODUCT INVENTORY REPORT"
            worksheet['A1'].font = Font(size=16, bold=True)
            worksheet['A1'].alignment = self.center_alignment
            worksheet.merge_cells('A1:H1')
            
            # Add generation date
            worksheet['A2'] = f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
            worksheet['A2'].font = Font(italic=True)
            worksheet.merge_cells('A2:H2')
            
            # Headers
            headers = ['Product ID', 'Name', 'SKU', 'Category', 'Price', 'Stock Quantity', 'Low Stock Alert', 'Status']
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=4, column=col, value=header)
            
            self._apply_header_style(worksheet, 4, 1, len(headers))
            
            # Data rows
            for row, product in enumerate(products_data, 5):
                worksheet.cell(row=row, column=1, value=product.get('id', ''))
                worksheet.cell(row=row, column=2, value=product.get('name', ''))
                worksheet.cell(row=row, column=3, value=product.get('sku', ''))
                worksheet.cell(row=row, column=4, value=product.get('category', ''))
                
                # Format price as currency
                price_cell = worksheet.cell(row=row, column=5, value=product.get('price', 0))
                price_cell.number_format = '₦#,##0.00'
                price_cell.alignment = self.right_alignment
                
                # Stock quantity
                stock_cell = worksheet.cell(row=row, column=6, value=product.get('stock_quantity', 0))
                stock_cell.alignment = self.right_alignment
                
                # Low stock alert threshold
                alert_cell = worksheet.cell(row=row, column=7, value=product.get('low_stock_alert', 0))
                alert_cell.alignment = self.right_alignment
                
                worksheet.cell(row=row, column=8, value=product.get('status', '').title())
                
                # Highlight low stock items
                if product.get('stock_quantity', 0) <= product.get('low_stock_alert', 0):
                    for col in range(1, len(headers) + 1):
                        cell = worksheet.cell(row=row, column=col)
                        cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
            
            # Apply data styling
            if products_data:
                self._apply_data_style(worksheet, 5, 4 + len(products_data), 1, len(headers))
            
            # Summary section
            summary_row = 6 + len(products_data)
            worksheet.cell(row=summary_row, column=1, value="SUMMARY").font = Font(bold=True)
            
            total_products = len(products_data)
            active_products = len([p for p in products_data if p.get('status') == 'active'])
            low_stock_products = len([p for p in products_data if p.get('stock_quantity', 0) <= p.get('low_stock_alert', 0)])
            
            worksheet.cell(row=summary_row + 1, column=1, value="Total Products:")
            worksheet.cell(row=summary_row + 1, column=2, value=total_products)
            
            worksheet.cell(row=summary_row + 2, column=1, value="Active Products:")
            worksheet.cell(row=summary_row + 2, column=2, value=active_products)
            
            worksheet.cell(row=summary_row + 3, column=1, value="Low Stock Products:")
            worksheet.cell(row=summary_row + 3, column=2, value=low_stock_products)
            
            total_inventory_value = sum(p.get('price', 0) * p.get('stock_quantity', 0) for p in products_data)
            worksheet.cell(row=summary_row + 4, column=1, value="Total Inventory Value:")
            inventory_value_cell = worksheet.cell(row=summary_row + 4, column=2, value=total_inventory_value)
            inventory_value_cell.number_format = '₦#,##0.00'
            
            self._auto_adjust_columns(worksheet)
            workbook.save(output_path)
            
            logger.info(f"Product Excel export generated successfully: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Failed to generate product Excel export: {str(e)}")
            return False
    
    def export_financial_summary_to_excel(self, financial_data: Dict[str, Any], output_path: str) -> bool:
        """
        Export financial summary to Excel
        
        Args:
            financial_data: Dictionary containing financial information
            output_path: Path where Excel file will be saved
        
        Returns:
            Boolean indicating success
        """
        try:
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            workbook = Workbook()
            
            # Revenue Summary Sheet
            revenue_sheet = workbook.active
            revenue_sheet.title = "Revenue Summary"
            
            revenue_sheet['A1'] = "FINANCIAL SUMMARY REPORT"
            revenue_sheet['A1'].font = Font(size=16, bold=True)
            revenue_sheet['A1'].alignment = self.center_alignment
            revenue_sheet.merge_cells('A1:C1')
            
            revenue_sheet['A2'] = f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
            revenue_sheet['A2'].font = Font(italic=True)
            revenue_sheet.merge_cells('A2:C2')
            
            # Revenue data
            revenue_data = [
                ['Metric', 'Amount', 'Notes'],
                ['Total Revenue', financial_data.get('total_revenue', 0), 'All-time revenue'],
                ['This Month', financial_data.get('revenue_this_month', 0), 'Current month revenue'],
                ['Last Month', financial_data.get('revenue_last_month', 0), 'Previous month revenue'],
                ['Outstanding', financial_data.get('outstanding_amount', 0), 'Unpaid invoices'],
            ]
            
            for row, data in enumerate(revenue_data, 4):
                for col, value in enumerate(data, 1):
                    cell = revenue_sheet.cell(row=row, column=col, value=value)
                    if row == 4:  # Header row
                        cell.font = self.header_font
                        cell.fill = self.header_fill
                        cell.alignment = self.center_alignment
                    elif col == 2 and row > 4:  # Amount column
                        cell.number_format = '₦#,##0.00'
                        cell.alignment = self.right_alignment
                    cell.border = self.border
            
            self._auto_adjust_columns(revenue_sheet)
            workbook.save(output_path)
            
            logger.info(f"Financial summary Excel export generated successfully: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Failed to generate financial summary Excel export: {str(e)}")
            return False

# Create a singleton instance
excel_service = ExcelService()

